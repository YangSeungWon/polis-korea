#!/usr/bin/env python3
"""NEC '성별·연령대별 투표율(구시군별)' XLSX → 시도×성별×연령대 투표율 JSON.

소스: 공공데이터포털 중앙선관위 '제N대 선거 투표율 분석' fileData(ZIP) 중
  01_투표율분석결과(표본)/11-성별·연령대별 투표율(구시군별).xlsx
시트 = 시도(서울/부산/...). 행 = (지역[전체|구시군] × 성별[합계|남자|여자] × 지표[선거인수|투표자수|투표율]),
열 = 연령대(18세/19세/20-24/25-29/30-34/35-39/40-49/50-59/60-69/70-79/80세이상) + 합계.

이 도구는 시도 '전체' 행만 뽑아(구시군은 후속) 연령대를 폴 크로스탭과 같은 띠로 집계:
  18-29 / 30 / 40 / 50 / 60 / 70+. 선거인수·투표자수에서 투표율 재계산(반올림 손실 회피).
전국 = 17 시도 합산.

출력: data/polls/turnout_demographics_<id>.json
  { "_meta":{...}, "전국":{"합계":{"18-29":{electors,voters,turnout}, ...},"남자":..,"여자":..},
    "서울특별시":{...}, ... }
"""
from __future__ import annotations
import json
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]

SHEET_SIDO = {
    "서울": "서울특별시", "부산": "부산광역시", "대구": "대구광역시", "인천": "인천광역시",
    "광주": "광주광역시", "대전": "대전광역시", "울산": "울산광역시", "세종": "세종특별자치시",
    "경기": "경기도", "강원": "강원특별자치도", "충북": "충청북도", "충남": "충청남도",
    "전북": "전북특별자치도", "전남": "전라남도", "경북": "경상북도", "경남": "경상남도",
    "제주": "제주특별자치도",
}
# 세분 연령열 → 폴 정렬 띠
AGE_BAND = {
    "18세": "18-29", "19세": "18-29", "20-24세": "18-29", "25-29세": "18-29",
    "30-34세": "30", "35-39세": "30", "40-49세": "40", "50-59세": "50",
    "60-69세": "60", "70-79세": "70+", "80세이상": "70+",
}
SEX_KEY = {"합계": "합계", "남자": "남자", "여자": "여자"}


def _num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def parse_sheet(ws):
    """시도 '전체' 블록 → {성별: {band: {electors, voters}}}."""
    rows = list(ws.iter_rows(values_only=True))
    # 헤더(연령열) 행 찾기
    hdr_i = next((i for i, r in enumerate(rows[:10])
                  if r and any(str(c).strip() == "18세" for c in r if c)), None)
    if hdr_i is None:
        return None
    header = [str(c).strip() if c is not None else "" for c in rows[hdr_i]]
    # 연령열 인덱스
    age_cols = [(j, AGE_BAND[h]) for j, h in enumerate(header) if h in AGE_BAND]
    if not age_cols:
        return None
    out = {"합계": {}, "남자": {}, "여자": {}}
    region = sex = metric = None
    for r in rows[hdr_i + 1:]:
        if not r:
            continue
        c0 = (str(r[0]).strip() if r[0] is not None else "")
        c1 = (str(r[1]).strip() if r[1] is not None else "")
        c2 = (str(r[2]).strip() if r[2] is not None else "")
        if c0:
            region = c0
        if c1 in SEX_KEY:
            sex = SEX_KEY[c1]
        if c2:
            metric = c2
        if region != "전체" or sex is None or metric not in ("선거인수", "투표자수"):
            continue
        field = "electors" if metric == "선거인수" else "voters"
        for j, band in age_cols:
            v = _num(r[j]) if j < len(r) else None
            if v is None:
                continue
            d = out[sex].setdefault(band, {"electors": 0.0, "voters": 0.0})
            d[field] += v
    return out


def finalize(blocks: dict):
    """electors/voters → turnout % 추가. 빈 시도 제거."""
    for sido, sexes in blocks.items():
        for sex, bands in sexes.items():
            for band, d in bands.items():
                e, v = d.get("electors", 0), d.get("voters", 0)
                d["electors"] = int(e)
                d["voters"] = int(v)
                d["turnout"] = round(v / e * 100, 1) if e else None
    return blocks


def main():
    if len(sys.argv) < 3:
        print("usage: parse_turnout_demographics.py <xlsx> <election_id>", file=sys.stderr)
        sys.exit(2)
    xlsx, eid = Path(sys.argv[1]), sys.argv[2]
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    blocks = {}
    for sh in wb.sheetnames:
        sido = SHEET_SIDO.get(sh.strip())
        if not sido:
            continue
        parsed = parse_sheet(wb[sh])
        if parsed:
            blocks[sido] = parsed
    # 전국 = 시도 합산
    nat = {"합계": {}, "남자": {}, "여자": {}}
    for sido, sexes in blocks.items():
        for sex, bands in sexes.items():
            for band, d in bands.items():
                t = nat[sex].setdefault(band, {"electors": 0.0, "voters": 0.0})
                t["electors"] += d.get("electors", 0)
                t["voters"] += d.get("voters", 0)
    result = {"_meta": {"election": eid, "source": "NEC 투표율분석(표본) 성별·연령대별",
                        "note": "선거인수·투표자수 합산 후 투표율 재계산. 연령 띠: 18-29/30/40/50/60/70+"},
              "전국": finalize({"_": nat})["_"]}
    result.update(finalize(blocks))
    dst = ROOT / f"data/polls/turnout_demographics_{eid.split('-')[0].replace('st','').replace('th','').replace('nd','').replace('rd','')}pres.json"
    dst.write_text(json.dumps(result, ensure_ascii=False, separators=(",", ":")))
    # 요약
    n = result["전국"]
    print(f"{eid}: {len(blocks)} 시도 + 전국 → {dst.name}")
    print("전국 투표율(합계):", {b: n["합계"][b]["turnout"] for b in ["18-29", "30", "40", "50", "60", "70+"] if b in n["합계"]})
    print("전국 남자:", {b: n["남자"][b]["turnout"] for b in ["18-29", "60", "70+"] if b in n["남자"]})
    print("전국 여자:", {b: n["여자"][b]["turnout"] for b in ["18-29", "60", "70+"] if b in n["여자"]})


if __name__ == "__main__":
    main()
