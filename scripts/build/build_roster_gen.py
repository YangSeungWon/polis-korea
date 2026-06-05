"""22대 총선 후보 명부(roster) 생성 — nec_district/n22_district_nec.xlsx → nec_roster_22gen.json.

build_polls_gen.py(향후)가 지역구 후보 race를 anchor하고 정당을 backfill하는 데 사용.
대선처럼 후보를 하드코딩할 수 없어(254 선거구 ≈ 700 후보) NEC 개표결과 xlsx에서 추출.

xlsx 구조 ('지역구' 시트): 선거구마다 머리 2행 —
  · 정당행: 읍면동 칸 빈 채로 col6+ 에 정당명 나열 (더불어민주당·국민의힘·…)
  · 후보행: 바로 다음 행, 같은 칸에 후보명 나열 (곽상언·최재형·…)
이후 합계/읍면동별 득표 데이터행. 두 머리행을 짝지어 {후보명: 정당} 추출.

출력 스키마:
  {
    "_meta": {...},
    "districts": { "시도|선거구": {"후보명": "정당", ...}, ... },   # 선거구명은 xlsx 원형(공백X)
    "name_party": { "후보명": "정당", ... },                       # 전역 fallback (동명이인 시 다수결)
    "proportional_parties": ["더불어민주연합", "국민의미래", ...]    # 비례 정당투표 검증용
  }

선거구명은 xlsx 원형(예: "중구성동구갑", "동두천시양주시연천군갑") 그대로. NESDC region은
공백·"선거구" 포함("부산광역시 부산진구 갑 선거구")이라 build 쪽에서 정규화해 매칭.
"""

from __future__ import annotations
import json
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
XLSX = ROOT / "data" / "raw" / "nec_district" / "n22_district_nec.xlsx"
OUT = ROOT / "data" / "raw" / "nec_roster_22gen.json"


def build_districts(ws) -> dict[str, dict[str, str]]:
    """선거구별 {후보명: 정당} 추출."""
    # (시도, 선거구) 그룹 순서 유지하며, 읍면동 칸 빈 머리행 2개(정당/후보)만 수집
    groups: dict[tuple[str, str], list[tuple]] = defaultdict(list)
    for r in ws.iter_rows(min_row=2, values_only=True):
        sido, district, eupmyeon = r[0], r[1], r[2]
        if not sido or not district:
            continue
        if eupmyeon in (None, "", " "):  # 머리행 (정당행/후보행)
            groups[(sido, district)].append(r)

    out: dict[str, dict[str, str]] = {}
    bad = []
    for (sido, district), heads in groups.items():
        if len(heads) < 2:
            bad.append((sido, district, len(heads)))
            continue
        party_row, name_row = heads[0], heads[1]
        # 후보 컬럼 = 정당행 셀이 비지 않은 인덱스 (col6부터 '계' 직전까지)
        mapping: dict[str, str] = {}
        for i, party in enumerate(party_row):
            if i < 6 or not party or not isinstance(party, str):
                continue
            party = party.strip()
            name = name_row[i] if i < len(name_row) else None
            if not name or not isinstance(name, str):
                continue
            name = name.strip()
            if name:
                mapping[name] = party
        if mapping:
            out[f"{sido}|{district}"] = mapping
        else:
            bad.append((sido, district, 0))
    if bad:
        print(f"  ! 머리행 추출 실패 {len(bad)}건: {bad[:5]}", file=sys.stderr)
    return out


def extract_proportional_parties(ws) -> list[str]:
    """비례대표 시트 머리행에서 정당명 리스트."""
    for r in ws.iter_rows(min_row=2, max_row=2, values_only=True):
        return [c.strip() for c in r[6:] if c and isinstance(c, str)]
    return []


ELECTION_BY_N = {"22": "22nd-general-2024", "21": "21st-general-2020", "20": "20th-general-2016"}


def main():
    import argparse
    ap = argparse.ArgumentParser(description="총선 선거구 후보 roster 생성")
    ap.add_argument("--n", default="22", help="총선 회차 (22/21/20)")
    args = ap.parse_args()
    xlsx = ROOT / "data" / "raw" / "nec_district" / f"n{args.n}_district_nec.xlsx"
    out_path = ROOT / "data" / "raw" / f"nec_roster_{args.n}gen.json"
    election = ELECTION_BY_N.get(args.n, f"{args.n}th-general")

    wb = openpyxl.load_workbook(xlsx, read_only=True)
    districts = build_districts(wb["지역구"])
    prop_parties = extract_proportional_parties(wb["비례대표"])
    wb.close()

    # 전역 name→party (동명이인 다수결). 같은 이름 다른 정당이면 최빈값.
    name_votes: dict[str, Counter] = defaultdict(Counter)
    for cand_map in districts.values():
        for name, party in cand_map.items():
            name_votes[name][party] += 1
    name_party = {n: c.most_common(1)[0][0] for n, c in name_votes.items()}

    n_cands = sum(len(v) for v in districts.values())
    out = {
        "_meta": {
            "election": election,
            "source": f"NEC 개표결과 n{args.n}_district_nec.xlsx (data/raw/nec_district)",
            "n_districts": len(districts),
            "n_candidates": n_cands,
            "n_proportional_parties": len(prop_parties),
            "note": "선거구명은 xlsx 원형(공백 제거). build_polls_gen.py가 NESDC region을 정규화해 매칭.",
        },
        "districts": districts,
        "name_party": name_party,
        "proportional_parties": prop_parties,
    }
    out_path.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[{election}] 선거구 {len(districts)}, 후보 {n_cands}, 비례정당 {len(prop_parties)}", file=sys.stderr)
    print(f"→ {out_path.relative_to(ROOT)}", file=sys.stderr)


if __name__ == "__main__":
    main()
