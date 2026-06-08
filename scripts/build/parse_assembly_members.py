"""국회 OpenAPI '역대 국회의원 현황' xlsx → 의원 unique ID 매핑.

소스: data/raw/assembly/데이터_역대 국회의원 현황.xlsx (5,689 rows)
각 row = 한 인물. 한자명 + 생년월일 = unique key. '대별 및 소속정당' 텍스트에서
"제N대(지역구) 정당" 패턴 추출 → 회차별 출마 이력.

Output: data/raw/assembly_member_map.json
  {
    "persons": [
      {"id": "강기윤_1960-06-04", "name": "강기윤", "hanja": "姜起潤", "dob": "1960-06-04",
       "careers": [{"n": 19, "district": "경남 창원시성산구", "party": "새누리당"},
                   {"n": 21, "district": "경남 창원시성산구", "party": "미래통합당"}]},
      ...
    ]
  }

build_person_index.py가 (eid, name) → assembly_id 매핑에 사용.
"""
from __future__ import annotations
import json
import re
import warnings
from pathlib import Path

import openpyxl

warnings.filterwarnings("ignore")

ROOT = Path(__file__).resolve().parents[2]
XLSX_HISTORICAL = ROOT / "data/raw/assembly/데이터_역대 국회의원 현황.xlsx"
XLSX_CURRENT = ROOT / "data/raw/assembly/데이터_국회의원 의원이력.xlsx"
XLSX_PROFILE = ROOT / "data/raw/assembly/데이터_국회의원 인적사항.xlsx"
OUT = ROOT / "data/raw/assembly_member_map.json"

# 현황: "제N대[국회의원] (지역구) [정당]"
CAREER_RE = re.compile(
    r"제(\d+)대\s*(?:국회의원)?\s*\(([^)]+)\)\s*([^\s제]+(?:\s*[^\s제(]+)?)?",
    re.MULTILINE,
)
# 의원이력: "제N대 정당 지역구"
CAREER_RE_CURRENT = re.compile(r"제(\d+)대\s+(\S+)\s+(.+?)$")

def parse_dob(s: str) -> str | None:
    """'1960년 06월 04일' → '1960-06-04'."""
    if not s:
        return None
    m = re.match(r"(\d{4})\s*년\s*(\d{1,2})\s*월\s*(\d{1,2})", s)
    if not m:
        return None
    return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"


def parse_career(text: str) -> list[dict]:
    if not text:
        return []
    out = []
    for m in CAREER_RE.finditer(text):
        n = int(m.group(1))
        district = (m.group(2) or "").strip()
        party = (m.group(3) or "").strip() if m.group(3) else ""
        # party가 잘못 잡혔을 경우 (다음 "제N대"의 앞부분일 수도) — 검증
        if party.startswith("제") and "대" in party[:4]:
            party = ""
        out.append({"n": n, "district": district, "party": party})
    return out


def parse_career_current(text: str) -> dict | None:
    """의원이력 row의 '의원이력' 셀: '제22대 국민의힘 대구 동구군위군을' → 1건."""
    if not text:
        return None
    m = CAREER_RE_CURRENT.match(text.strip())
    if not m:
        return None
    return {"n": int(m.group(1)), "district": m.group(3).strip(), "party": m.group(2).strip()}


def main():
    by_key = {}
    n_merged = 0
    n_force_split = 0

    def merge_or_add(name, hanja, dob, careers):
        nonlocal n_merged, n_force_split
        if not careers:
            return
        if dob:
            key = f"{name}_{dob}"
        elif hanja:
            key = f"{name}_{hanja}"
        else:
            n_force_split += 1
            key = f"{name}_unk_{len(by_key)}"
        if key in by_key:
            existing = by_key[key]
            ek = {(c["n"], c["district"]) for c in existing["careers"]}
            for c in careers:
                if (c["n"], c["district"]) not in ek:
                    existing["careers"].append(c)
                    ek.add((c["n"], c["district"]))
            # hanja·dob 누락분 보완
            if not existing.get("hanja") and hanja:
                existing["hanja"] = hanja
            if not existing.get("dob") and dob:
                existing["dob"] = dob
            n_merged += 1
            return
        by_key[key] = {
            "id": key, "name": name, "hanja": hanja, "dob": dob, "careers": careers,
        }

    # 1. 역대 현황
    wb = openpyxl.load_workbook(XLSX_HISTORICAL)
    ws = wb.active
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0 or not row or not row[2]:
            continue
        merge_or_add(
            (row[2] or "").strip(),
            (row[3] or "").strip(),
            parse_dob(row[6] or ""),
            parse_career(row[1] or ""),
        )

    # 2. 의원이력 (현역 22대 + 직전 회차 보강)
    # dob 없는 entry라 단순 merge 시 새 키 생성 위험. 먼저 hanja로 existing 검색·merge.
    hanja_to_existing = {}
    for k, v in by_key.items():
        if v.get("hanja"):
            hanja_to_existing.setdefault((v["name"], v["hanja"]), []).append(v)

    wb = openpyxl.load_workbook(XLSX_CURRENT)
    ws = wb.active
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0 or not row or not row[1]:
            continue
        name = (row[1] or "").strip()
        hanja = (row[2] or "").strip()
        career = parse_career_current(row[4] or "")
        if not career:
            continue
        # 동일 한자명 entry 1건 → 그쪽에 careers 추가
        matches = hanja_to_existing.get((name, hanja), [])
        if len(matches) == 1:
            existing = matches[0]
            ek = {(c["n"], c["district"]) for c in existing["careers"]}
            if (career["n"], career["district"]) not in ek:
                existing["careers"].append(career)
            n_merged += 1
            continue
        # 매칭 없음 → 새 entry (hanja-based key)
        merge_or_add(name, hanja, None, [career])
    # 3. 인적사항 (22대 현역) — dob 보강
    wb = openpyxl.load_workbook(XLSX_PROFILE)
    ws = wb.active
    n_dob_filled = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0 or not row or not row[0]:
            continue
        name = (row[0] or "").strip()
        hanja = (row[1] or "").strip()
        dob_raw = row[4]
        if not dob_raw:
            continue
        # YYYY-MM-DD or datetime
        if hasattr(dob_raw, "strftime"):
            dob = dob_raw.strftime("%Y-%m-%d")
        else:
            dob = str(dob_raw).strip()[:10]
        # 기존 entry 찾기 — hanja 매칭 우선
        targets = []
        for k, v in by_key.items():
            if v["name"] == name and v.get("hanja") == hanja:
                targets.append(v)
        if not targets:
            for k, v in by_key.items():
                if v["name"] == name and not v.get("dob"):
                    targets.append(v)
        # dob 부재 entry만 보강 (이미 있으면 안 건드림)
        for t in targets:
            if not t.get("dob"):
                t["dob"] = dob
                # key 재배정 — name_hanja → name_dob
                old_id = t["id"]
                new_id = f"{name}_{dob}"
                if new_id != old_id and new_id not in by_key:
                    t["id"] = new_id
                    by_key[new_id] = t
                    by_key.pop(old_id, None)
                n_dob_filled += 1

    persons = list(by_key.values())
    # careers 회차 정렬
    for p in persons:
        p["careers"].sort(key=lambda c: c["n"])
    n_dup = n_merged

    # 정당명 캐노니컬 (옛 정당 보정 — 사람마다 표기 다를 수 있음)
    # 일단 raw 그대로 두고, build_person_index 단에서 매칭 처리.

    out = {
        "_meta": {
            "source": "open.assembly.go.kr 역대 국회의원 현황",
            "n_persons": len(persons),
            "n_careers": sum(len(p["careers"]) for p in persons),
        },
        "persons": persons,
    }
    OUT.write_text(json.dumps(out, ensure_ascii=False, indent=1), encoding="utf-8")
    print(f"→ {OUT.relative_to(ROOT)}: {len(persons)} persons · "
          f"{out['_meta']['n_careers']} careers · {OUT.stat().st_size/1024:.1f} KB")
    if n_dup:
        print(f"  dup keys auto-suffixed: {n_dup}")

    # 회차별 의원 수
    from collections import Counter
    by_n = Counter()
    for p in persons:
        for c in p["careers"]:
            by_n[c["n"]] += 1
    print("\n회차별 cumulative 의원 row:")
    for n in sorted(by_n):
        print(f"  제{n}대: {by_n[n]}")


if __name__ == "__main__":
    main()
